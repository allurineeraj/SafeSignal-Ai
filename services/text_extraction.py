import os
import fitz  # PyMuPDF
import docx
import openpyxl

def is_ocr_available() -> bool:
    """Check for OCR Tesseract binaries."""
    # We fallback to warning as per plan
    return False

def extract_text_from_file(file_path: str) -> str:
    """Parse PDF, Word, Excel, TXT or image text."""
    if not os.path.exists(file_path):
        return ""
        
    ext = file_path.lower().split('.')[-1]
    extracted_text = ""
    
    try:
        if ext == 'pdf':
            doc = fitz.open(file_path)
            for page in doc:
                extracted_text += page.get_text() + "\n"
            doc.close()
        elif ext == 'docx':
            doc = docx.Document(file_path)
            extracted_text = "\n".join([para.text for para in doc.paragraphs])
        elif ext == 'txt':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                extracted_text = f.read()
        elif ext == 'xlsx':
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        extracted_text += row_text + "\n"
            wb.close()
        elif ext in ['csv']:
            import pandas as pd
            df = pd.read_csv(file_path)
            extracted_text = df.to_string(index=False)
        elif ext in ['png', 'jpg', 'jpeg']:
            extracted_text = "[Image attached. OCR not configured. Manual review required.]"
        elif ext in ['wav', 'mp3']:
            extracted_text = "[Audio attached. Transcription unavailable in fallback mode.]"
        else:
            extracted_text = "[Unsupported file format for text extraction.]"
    except Exception as e:
        extracted_text = f"[Error extracting text: {str(e)}]"
        
    return extracted_text.strip()
